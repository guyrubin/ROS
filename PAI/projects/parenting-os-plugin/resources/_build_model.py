from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = Font(name='Arial', size=10, color='0000FF')      # inputs
BLACK = Font(name='Arial', size=10, color='000000')     # formulas
GREEN = Font(name='Arial', size=10, color='008000')     # cross-sheet links
BOLD = Font(name='Arial', size=10, bold=True, color='000000')
TITLE = Font(name='Arial', size=14, bold=True, color='1F4E2C')
HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')
SUB = Font(name='Arial', size=11, bold=True, color='1F4E2C')
NOTE = Font(name='Arial', size=9, italic=True, color='666666')
HDR_FILL = PatternFill('solid', fgColor='1F4E2C')
SEC_FILL = PatternFill('solid', fgColor='E3EFE6')
YEL = PatternFill('solid', fgColor='FFFDE7')
thin = Side(style='thin', color='CCCCCC')
TOPB = Border(top=Side(style='thin', color='888888'))

EUR = u'€#,##0;(€#,##0);"-"'
EUR2 = u'€#,##0.00;(€#,##0.00);"-"'
NUM = '#,##0;(#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'

wb = Workbook()

# ---------------- ASSUMPTIONS ----------------
a = wb.active
a.title = 'Assumptions'
a.sheet_view.showGridLines = False
a['A1'] = 'Arbor — Financial Model'; a['A1'].font = TITLE
a['A2'] = 'Base case. Blue cells = editable inputs. All currency in € thousands (€000) unless noted /mo.'
a['A2'].font = NOTE
a['A3'] = 'v1.0 · 2026-06-07 · PAI / Arbor · companion to arbor-business-model-pricing-and-financials.md'
a['A3'].font = NOTE

def yhead(ws, row):
    for col, lbl in zip('BCD', ['Y1 2026', 'Y2 2027', 'Y3 2028']):
        c = ws[f'{col}{row}']; c.value = lbl; c.font = HDR; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center')
    ws[f'A{row}'].fill = HDR_FILL

def section(ws, row, label):
    ws[f'A{row}'] = label; ws[f'A{row}'].font = SUB
    for col in 'ABCD':
        ws[f'{col}{row}'].fill = SEC_FILL

def inrow(ws, row, label, vals, fmt=NUM, note=None):
    ws[f'A{row}'] = label; ws[f'A{row}'].font = BLACK
    for col, v in zip('BCD', vals):
        c = ws[f'{col}{row}']; c.value = v; c.font = BLUE; c.number_format = fmt
        c.fill = YEL
    if note:
        ws[f'F{row}'] = note; ws[f'F{row}'].font = NOTE

yhead(a, 5)
section(a, 6, 'B2C — Consumer subscription')
inrow(a, 7, 'Avg paying subscribers', [1500, 9000, 38000], NUM)
inrow(a, 8, 'Blended ARPU (€/mo)', [11.0, 11.0, 11.5], EUR2)
inrow(a, 9, 'Fully-loaded COGS per paying user (€/mo)', [1.40, 1.50, 1.60], EUR2,
      'AI (Claude+Gemini) + infra + Stripe')
section(a, 10, 'Professional — seats')
inrow(a, 11, 'Avg active seats', [20, 140, 480], NUM)
inrow(a, 12, 'Price per seat (€/mo)', [45, 45, 45], EUR)
inrow(a, 13, 'COGS per seat (€/mo)', [4, 4, 4], EUR)
section(a, 14, 'Institutional — B2B2C / B2G')
inrow(a, 15, 'Avg active subsidised families', [0, 1500, 12000], NUM)
inrow(a, 16, 'ARPF (€/mo)', [0, 5.0, 5.5], EUR2)
inrow(a, 17, 'Pilot & setup fees (€000/yr)', [30, 30, 0], EUR, 'Fixed-fee cohort pilots')
inrow(a, 18, 'Institutional COGS (% of recurring)', [0.15, 0.15, 0.15], PCT)
section(a, 19, 'Operating costs (€000)')
inrow(a, 20, 'Team (loaded, NL/IL blend)', [295, 750, 1600], EUR)
inrow(a, 21, 'Sales & marketing', [20, 180, 500], EUR)
inrow(a, 22, 'Compliance / security / legal', [30, 60, 120], EUR)
inrow(a, 23, 'Content / expert review / localisation', [25, 60, 150], EUR)
inrow(a, 24, 'G&A / tooling / SaaS', [15, 60, 130], EUR)

a.column_dimensions['A'].width = 42
for col in 'BCD':
    a.column_dimensions[col].width = 13
a.column_dimensions['F'].width = 30

# ---------------- MODEL ----------------
m = wb.create_sheet('Model')
m.sheet_view.showGridLines = False
m['A1'] = 'Arbor — 3-Year P&L (€000)'; m['A1'].font = TITLE
m['A2'] = 'Black = formula · Green = link to Assumptions. Re-forecast monthly off real analytics.'
m['A2'].font = NOTE
yhead(m, 4)

def lbl(ws, row, text, font=BLACK):
    ws[f'A{row}'] = text; ws[f'A{row}'].font = font

def frow(ws, row, formulas, fmt=EUR, font=BLACK, fill=None, border=False):
    for col, f in zip('BCD', formulas):
        c = ws[f'{col}{row}']; c.value = f; c.font = font; c.number_format = fmt
        if fill: c.fill = fill
        if border: c.border = TOPB

cols = ['B', 'C', 'D']
# Revenue
section(m, 5, 'Revenue')
lbl(m, 6, 'B2C subscription')
frow(m, 6, [f'=Assumptions!{c}7*Assumptions!{c}8*12/1000' for c in cols], EUR, GREEN)
lbl(m, 7, 'Professional seats')
frow(m, 7, [f'=Assumptions!{c}11*Assumptions!{c}12*12/1000' for c in cols], EUR, GREEN)
lbl(m, 8, 'Institutional recurring')
frow(m, 8, [f'=Assumptions!{c}15*Assumptions!{c}16*12/1000' for c in cols], EUR, GREEN)
lbl(m, 9, 'Institutional pilot/setup fees')
frow(m, 9, [f'=Assumptions!{c}17' for c in cols], EUR, GREEN)
lbl(m, 10, 'Total revenue', BOLD)
frow(m, 10, [f'=SUM({c}6:{c}9)' for c in cols], EUR, BOLD, border=True)

# COGS
section(m, 12, 'Cost of goods sold')
lbl(m, 13, 'B2C COGS')
frow(m, 13, [f'=Assumptions!{c}7*Assumptions!{c}9*12/1000' for c in cols], EUR, GREEN)
lbl(m, 14, 'Professional COGS')
frow(m, 14, [f'=Assumptions!{c}11*Assumptions!{c}13*12/1000' for c in cols], EUR, GREEN)
lbl(m, 15, 'Institutional COGS')
frow(m, 15, [f'={c}8*Assumptions!{c}18' for c in cols], EUR, BLACK)
lbl(m, 16, 'Total COGS', BOLD)
frow(m, 16, [f'=SUM({c}13:{c}15)' for c in cols], EUR, BOLD, border=True)
lbl(m, 17, 'Gross profit', BOLD)
frow(m, 17, [f'={c}10-{c}16' for c in cols], EUR, BOLD)
lbl(m, 18, 'Gross margin %')
frow(m, 18, [f'={c}17/{c}10' for c in cols], PCT, BLACK)

# OpEx
section(m, 20, 'Operating expenses')
lbl(m, 21, 'Team')
frow(m, 21, [f'=Assumptions!{c}20' for c in cols], EUR, GREEN)
lbl(m, 22, 'Sales & marketing')
frow(m, 22, [f'=Assumptions!{c}21' for c in cols], EUR, GREEN)
lbl(m, 23, 'Compliance / security / legal')
frow(m, 23, [f'=Assumptions!{c}22' for c in cols], EUR, GREEN)
lbl(m, 24, 'Content / expert / localisation')
frow(m, 24, [f'=Assumptions!{c}23' for c in cols], EUR, GREEN)
lbl(m, 25, 'G&A / tooling')
frow(m, 25, [f'=Assumptions!{c}24' for c in cols], EUR, GREEN)
lbl(m, 26, 'Total OpEx', BOLD)
frow(m, 26, [f'=SUM({c}21:{c}25)' for c in cols], EUR, BOLD, border=True)

# EBITDA
section(m, 28, 'Profitability')
lbl(m, 29, 'EBITDA', BOLD)
frow(m, 29, [f'={c}17-{c}26' for c in cols], EUR, BOLD)
lbl(m, 30, 'EBITDA margin %')
frow(m, 30, [f'={c}29/{c}10' for c in cols], PCT, BLACK)
lbl(m, 31, 'Cumulative EBITDA (cash proxy)', BOLD)
m['B31'] = '=B29'; m['C31'] = '=B31+C29'; m['D31'] = '=C31+D29'
for c in cols:
    m[f'{c}31'].font = BOLD; m[f'{c}31'].number_format = EUR
lbl(m, 33, 'Peak cumulative cash need (€000)')
m['B33'] = '=-MIN(B31:D31)'; m['B33'].font = BLACK; m['B33'].number_format = EUR
m['D33'] = 'Seed sizing carries buffer to €2.0–2.5M for slower institutional cycles.'
m['D33'].font = NOTE

m.column_dimensions['A'].width = 34
for col in 'BCD':
    m.column_dimensions[col].width = 13

# ---------------- SENSITIVITY ----------------
s = wb.create_sheet('Sensitivity')
s.sheet_view.showGridLines = False
s['A1'] = 'Arbor — Year-3 Scenario Sensitivity (€000)'; s['A1'].font = TITLE
s['A2'] = 'Illustrative. Factor scales Y3 base revenue & COGS; OpEx held semi-fixed. Blue = input.'
s['A2'].font = NOTE
for col, lblh in zip('BCD', ['Conservative', 'Base', 'Aggressive']):
    c = s[f'{col}4']; c.value = lblh; c.font = HDR; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
s['A4'].fill = HDR_FILL
s['A5'] = 'Revenue factor vs base'; s['A5'].font = BLACK
for col, v in zip('BCD', [0.62, 1.00, 1.50]):
    c = s[f'{col}5']; c.value = v; c.font = BLUE; c.number_format = '0.00x'; c.fill = YEL
s['A6'] = 'Lever (driver)'; s['A6'].font = BLACK
for col, v in zip('BCD', ['conv 4% / churn 7%', 'conv 6% / churn 5.5%', 'conv 8% + 2 muni deals']):
    s[f'{col}6'] = v; s[f'{col}6'].font = NOTE; s[f'{col}6'].alignment = Alignment(horizontal='center')
s['A7'] = 'Y3 revenue'; s['A7'].font = BLACK
frow_cols = 'BCD'
for col in frow_cols:
    s[f'{col}7'] = f'=Model!D10*{col}5'; s[f'{col}7'].font = GREEN; s[f'{col}7'].number_format = EUR
s['A8'] = 'Y3 COGS'; s['A8'].font = BLACK
for col in frow_cols:
    s[f'{col}8'] = f'=Model!D16*{col}5'; s[f'{col}8'].font = GREEN; s[f'{col}8'].number_format = EUR
s['A9'] = 'Y3 OpEx (semi-fixed)'; s['A9'].font = BLACK
for col in frow_cols:
    s[f'{col}9'] = '=Model!D26'; s[f'{col}9'].font = GREEN; s[f'{col}9'].number_format = EUR
s['A10'] = 'Y3 EBITDA'; s['A10'].font = BOLD
for col in frow_cols:
    s[f'{col}10'] = f'={col}7-{col}8-{col}9'; s[f'{col}10'].font = BOLD; s[f'{col}10'].number_format = EUR
    s[f'{col}10'].border = TOPB
s['A11'] = 'Y3 EBITDA margin'; s['A11'].font = BLACK
for col in frow_cols:
    s[f'{col}11'] = f'={col}10/{col}7'; s[f'{col}11'].font = BLACK; s[f'{col}11'].number_format = PCT
s.column_dimensions['A'].width = 24
for col in 'BCD':
    s.column_dimensions[col].width = 18

import os
out = os.path.join(os.path.dirname(__file__), 'arbor-financial-model.xlsx')
wb.save(out)
print('saved', out)
